# DESARROLLADO POR NICOLÁS ACUÑA

import sys
import os
import io
from datetime import *
import time as tiempo

# Librerías para exportar a excel
import pandas as pd
import xlsxwriter as xls
from openpyxl import load_workbook

# Librerías Interfaz Gráfica Tkinter
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkcalendar import dateentry
from tkinter import font


# Librería conexión base de datos por medio de mysql
import pymysql

# Librerías edición docx
from pathlib import Path
from docxtpl import DocxTemplate
from docx2pdf import convert


# Librerías impresión
from win32 import win32print
from win32 import win32api


fecha_anterior=datetime.now().date().strftime('%d-%m-%Y')


# LISTAS ASUNTO
lista_asunto_ingreso=[
    'ARRIENDO',
    'ARRIENDO EVENTO',
    'ARRIENDO FIN DE SEMANA',
    'PASE JUGADOR',
    'PRÉSTAMO JUGADOR',
    'BORDERÓ',
    'APORTE',
    'PUBLICIDAD',
    'FONDOS CONCURSABLES',
    'APORTE PARTICULAR',
    'OTRO']
lista_asunto_egreso=[
    'COPA PANCHO',
    'AGUA',
    'REMUNERACIONES',
    'HONORARIOS',
    'IMPOSICIONES',
    'IMPUESTO MENSUAL',
    'CONTRIBUCIONES',
    'CELULAR',
    'CAJA CHICA',
    'MANTENCIÓN GENERAL',
    'MANTENCIÓN ELECTRÓNICA',
    'ARBITROS',
    'CRONOMETRISTA',
    'DEUDA CONVENIO',
    'GASTOS COPA PANCHO',
    'ELECTRICIDAD',
    'IMPRENTA',
    'TRABAJO EXTRA',
    'TRANSPORTE',
    'COMUNICACIONES',
    'AGUINALDO',
    'DIVSIONES FORMATIVAS',
    'VARIOS',
    'OTRO'
]

# LISTAS RECIBIDO DE/ENVIADO A
lista_recibido_de=[
    'CONFITERÍA EL FORTÍN',
    'CONFITERÍA ACHAVAR',
    'LIN DÍAZ',
    'LIN DÍAZ SEGUNDO PISO',
    'LA VERTIENTE (S BERNAL)',
    'KOPPA (E BERNAL)',
    'CLAUDIA (E BERNAL)',
    'CAFETERIA (BRAULIO)',
    'SCOUT',
    'LIN DIAZ',
    'SAMUEL BERNAL',
    'ERIKA BERNAL',
    'SEÑOR PANELLI',
    'MUNICIPALIDAD',
    'PUERTO VALPARAISO',
    'PARTICULAR',
    'OTRO']
lista_enviado_a=[
    'PREVIRED',
    'TESORERÍA GENERAL DE LA REPÚBLICA',
    'CHILQUINTA',
    'GASVALPO',
    'ESVAL',
    'PERSONAL',
    'ASOCIACIÓN',
    'SECRETARIA',
    'CLAUDIO OSORIO',
    'VERONICA SAMIT',
    'BRAULIO SASSO',
    'MYRIAM VERDUGO',
    'ARNALDO PANELLI',
    'DELIA PANELLI',
    'LUIS SALINAS',
    'JAIME NEGRON',
    'JAVIER ARENAS',
    'PEDRO MELLA',
    'JOSE GUERRA',
    'BANCO SANTANDER',
    'ANA MARIA GRANADA ALVAREZ',
    'LUIS SANCHEZ',
    'OTRO'
]

# lista de meses
lista_meses=[
    '01 - ENERO',
    '02 - FEBRERO',
    '03 - MARZO',
    '04 - ABRIL',
    '05 - MAYO',
    '06 - JUNIO',
    '07 - JULIO',
    '08 - AGOSTO',
    '09 - SEPTIEMBRE',
    '10 - OCTUBRE',
    '11 - NOVIEMBRE',
    '12 - DICIEMBRE']

# lista de años
lista_a=[]
for a in range(2015,2031):
    lista_a.append(str(a))

#------------ SECCIÓN FUNCIÓN MONTO EN PALABRAS -------------------------------
MAX_NUMERO = 999999999999

UNIDADES = (
    'CERO',
    'UNO',
    'DOS',
    'TRES',
    'CUATRO',
    'CINCO',
    'SEIS',
    'SIETE',
    'OCHO',
    'NUEVE'
)

DECENAS = (
    'DIEZ',
    'ONCE',
    'DOCE',
    'TRECE',
    'CATORCE',
    'QUINCE',
    'DIECISEIS',
    'DIECISIETE',
    'DIECIOCHO',
    'DIECINUEVE'
)

DIEZ_DIEZ = (
    'CERO',
    'DIEZ',
    'VEINTE',
    'TREINTA',
    'CUARENTA',
    'CINCUENTA',
    'SESENTA',
    'SETENTA',
    'OCHENTA',
    'NOVENTA'
)

CIENTOS = (
    '_',
    'CIENTO',
    'DOSCIENTOS',
    'TRESCIENTOS',
    'CUATROSCIENTOS',
    'QUINIENTOS',
    'SEISCIENTOS',
    'SETECIENTOS',
    'OCHOCIENTOS',
    'NOVECIENTOS'
)

class Formato:
    def __init__(self) -> None:
        pass

    def numero_a_moneda_sunat(self, numero):
        numero_entero = int(numero)
        letras = numero_a_letras(numero_entero)
        letras = letras.replace('UNO', 'UN')
        letras = f"{letras}"
        return letras

    
def numero_a_letras(numero):
    numero_entero = int(numero)
    if numero_entero > MAX_NUMERO:
        raise OverflowError('Número demasiado alto')
    if numero_entero < 0:
        negativo_letras = numero_a_letras(abs(numero))
        return f"MENOS {negativo_letras}"
    if numero_entero <= 99:
        resultado = leer_decenas(numero_entero)
    elif numero_entero <= 999:
        resultado = leer_centenas(numero_entero)
    elif numero_entero <= 999999:
        resultado = leer_miles(numero_entero)
    elif numero_entero <= 999999999:
        resultado = leer_millones(numero_entero)
    else:
        resultado = leer_millardos(numero_entero)
    resultado = resultado.replace('UNO MIL', 'UN MIL')
    resultado = resultado.strip()
    resultado = resultado.replace(' _ ', ' ')
    resultado = resultado.replace('  ', ' ')
    return resultado


def numero_a_moneda(numero):
    numero_entero = int(numero)
    letras = numero_a_letras(numero_entero)
    letras = letras.replace('UNO', 'UN')
    letras = f"{letras}"
    return letras


def leer_decenas(numero):
    if numero < 10:
        return UNIDADES[numero]
    decena, unidad = divmod(numero, 10)
    if numero <= 19:
        resultado = DECENAS[unidad]
    elif numero <= 29:
        resultado = f"VEINTI{UNIDADES[unidad]}"
    else:
        resultado = DIEZ_DIEZ[decena]
        if unidad > 0:
            resultado = f"{resultado} Y {UNIDADES[unidad]}"
    return resultado


def leer_centenas(numero):
    centena, decena = divmod(numero, 100)
    if numero == 0:
        resultado = 'CIEN'
    else:
        resultado = CIENTOS[centena]
        if decena > 0:
            decena_letras = leer_decenas(decena)
            resultado = f"{resultado} {decena_letras}"
    return resultado


def leer_miles(numero):
    millar, centena = divmod(numero, 1000)
    resultado = ''
    if millar == 1:
        resultado = ''
    if (millar >= 2) and (millar <= 9):
        resultado = UNIDADES[millar]
    elif (millar >= 10) and (millar <= 99):
        resultado = leer_decenas(millar)
    elif (millar >= 100) and (millar <= 999):
        resultado = leer_centenas(millar)
    resultado = f"{resultado} MIL"
    if centena > 0:
        centena_letras = leer_centenas(centena)
        resultado = f"{resultado} {centena_letras}"
    return resultado.strip()


def leer_millones(numero):
    millon, millar = divmod(numero, 1000000)
    resultado = ''
    if millon == 1:
        resultado = ' UN MILLON '
    if (millon >= 2) and (millon <= 9):
        resultado = UNIDADES[millon]
    elif (millon >= 10) and (millon <= 99):
        resultado = leer_decenas(millon)
    elif (millon >= 100) and (millon <= 999):
        resultado = leer_centenas(millon)
    if millon > 1:
        resultado = f"{resultado} MILLONES"
    if (millar > 0) and (millar <= 999):
        centena_letras = leer_centenas(millar)
        resultado = f"{resultado} {centena_letras}"
    elif (millar >= 1000) and (millar <= 999999):
        miles_letras = leer_miles(millar)
        resultado = f"{resultado} {miles_letras}"
    return resultado


def leer_millardos(numero):
    millardo, millon = divmod(numero, 1000000)
    miles_letras = leer_miles(millardo)
    millones_letras = leer_millones(millon)
    return f"{miles_letras} MILLONES {millones_letras}"

# DATOS DE LA BASE DE DATOS REMOTA
DB_HOST = 'us-east.connect.psdb.cloud' 
DB_USER = 'hpn9tpk1dry0lmzu7377' 
DB_USER_PASSWORD = 'pscale_pw_z8FWTgp8gcRZJUEmtlO1GTY2mep54VHI46hlq7lxyOm' 
DB_NAME = 'administracion-ingresos-egresos'

# Importación inicial de los datos de la base de datos
def importar_datos_baseDeDatos():
    try:
        conexion_bdd = pymysql.connect(user=DB_USER, password=DB_USER_PASSWORD, host=DB_HOST, database=DB_NAME, cursorclass=pymysql.cursors.DictCursor, ssl={"rejectUnauthorized":False})
        mycursor = conexion_bdd.cursor()
        query="SELECT * FROM Transaccion ORDER BY substring(`numero`, 5) ASC, substring(`numero`, 1, 3) ASC, `tipo` ASC"
        mycursor.execute(query) 
        fila = mycursor.fetchall()
        mycursor.close()
        conexion_bdd.close()

        # Insercción de los datos en la tabla
        for dato in fila:
            if dato['nCheque']!=0:
                tabla.insert('', 'end', values=(dato['numero'], dato['tipo'], dato['asunto'], dato['persona'], dato['fecha'].strftime("%d-%m-%Y"), dato['medio'], dato['nCheque'], '{:,}'.format(dato['monto']).replace(',','.'), dato['descripcion']))
            else: tabla.insert('', 'end', values=(dato['numero'], dato['tipo'], dato['asunto'], dato['persona'], dato['fecha'].strftime("%d-%m-%Y"), dato['medio'], "--------", '{:,}'.format(dato['monto']).replace(',','.'), dato['descripcion']))
    
    except pymysql.Error as error:
        if messagebox.showerror(title="Error", message=error):
            app.destroy()


# Obtención del número de folio para la funcionalidades de agregar ingreso y agregar egreso
def obtener_numeroDeFolio_baseDeDatos():
    try:
        conexion_bdd = pymysql.connect(user=DB_USER, password=DB_USER_PASSWORD, host=DB_HOST, database=DB_NAME, cursorclass=pymysql.cursors.DictCursor, ssl={"rejectUnauthorized":False})
        mycursor=conexion_bdd.cursor()
        mycursor.execute("SELECT COUNT(*) FROM Transaccion WHERE `numero` LIKE '%"+str(fecha.year)+"' AND `tipo` LIKE '"+str(tipoT)+"'") # Sentencia MYSQL: Se cuentan todos los ingresos o egresos de un mismo año
        fila = mycursor.fetchall()
        mycursor.close()
        conexion_bdd.close()

        global numero
        ne=int(fila[0]['count(*)'])
        if ne>=0 and ne<=9:
            numero = "00"+str(fila[0]['count(*)'])+"-"+str(fecha.year)
        elif ne>=10 and ne<=99:
            numero = "0"+str(fila[0]['count(*)'])+"-"+str(fecha.year)
        else:
            numero = str(fila[0]['count(*)'])+"-"+str(fecha.year)

    except pymysql.Error as error:
        if messagebox.showerror(title="Error", message=error):
            numero='000-0000'
            cerrar_seccion_agregar()

# Insercción de dato nuevo en la base de datos
def insertar_dato_baseDeDatos():
    try:
        conexion_bdd = pymysql.connect(user=DB_USER, password=DB_USER_PASSWORD, host=DB_HOST, database=DB_NAME, cursorclass=pymysql.cursors.DictCursor, ssl={"rejectUnauthorized":False})
        sql = "INSERT INTO Transaccion (numero, tipo, asunto, persona, fecha, medio, nCheque, monto, descripcion) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)" # Sentenica MYSQL: Se inserta la fila nueva con sus datos
        valores = (numero, tipo, asunto, persona, date.isoformat(fecha), medio, ncheque, monto, descripcion)

        mycursor = conexion_bdd.cursor()
        mycursor.execute(sql, valores)
        conexion_bdd.commit()

        mycursor.close()
        conexion_bdd.close()

        # El elemento nuevo se inserta en la tabla para mantener ésta actualizada
        if ncheque!=0:
            tabla.insert('', 'end', values=(numero, tipo, asunto, persona, fecha.strftime("%d-%m-%Y"), medio, ncheque, '{:,}'.format(monto).replace(',','.'), descripcion))
        else: tabla.insert('', 'end', values=(numero, tipo, asunto, persona, fecha.strftime("%d-%m-%Y"), medio, "--------", '{:,}'.format(monto).replace(',','.'), descripcion))

    except pymysql.Error as error:
        if messagebox.showerror(title="Error", message=error):
            cerrar_seccion_agregar()

# Limpieza (refresh) de la tabla
def limpiar_busqueda_baseDeDatos():
    try:
        conexion_bdd = pymysql.connect(user=DB_USER, password=DB_USER_PASSWORD, host=DB_HOST, database=DB_NAME, cursorclass=pymysql.cursors.DictCursor, ssl={"rejectUnauthorized":False})
        tabla.delete(*tabla.get_children()) # Se elimina el contenido de la tabla actual
        tabla.heading('4', text="Recibido de/Enviado a", anchor=W)
        mycursor = conexion_bdd.cursor()
        query="SELECT * FROM Transaccion ORDER BY substring(`numero`, 5) ASC, substring(`numero`, 1, 3) ASC, `tipo` ASC"
        mycursor.execute(query) 
        fila = mycursor.fetchall()
        mycursor.close()
        conexion_bdd.close()

        # Se insertan en la tabla todos los elementos de la base de datos
        for dato in fila:
            if dato['nCheque']!=0:
                tabla.insert('', 'end', values=(dato['numero'], dato['tipo'], dato['asunto'], dato['persona'], dato['fecha'].strftime("%d-%m-%Y"), dato['medio'], dato['nCheque'], '{:,}'.format(dato['monto']).replace(',','.'), dato['descripcion']))
            else: tabla.insert('', 'end', values=(dato['numero'], dato['tipo'], dato['asunto'], dato['persona'], dato['fecha'].strftime("%d-%m-%Y"), dato['medio'], "--------", '{:,}'.format(dato['monto']).replace(',','.'), dato['descripcion']))

    except pymysql.Error as error:
        messagebox.showerror(title="Error", message=error)


# Búsqueda de datos por asunto/Filtro por tipo, año y rango de números de folio
def buscar_filtrar_baseDeDatos():
    try:
        conexion_bdd = pymysql.connect(user=DB_USER, password=DB_USER_PASSWORD, host=DB_HOST, database=DB_NAME, cursorclass=pymysql.cursors.DictCursor, ssl={"rejectUnauthorized":False})
        tabla.delete(*tabla.get_children()) # Se elimina el contenido de la tabla actual
        mycursor = conexion_bdd.cursor()
        if busqueda_var.get()=='':
            if filtroTipo_var.get()!='Todos' and filtroA_var.get()!='Todos':
                mycursor.execute("SELECT * FROM Transaccion WHERE `tipo` = '"+filtroTipo_var.get()+"' AND YEAR(fecha) = '"+filtroA_var.get()+"' AND substring(`numero`, 1, 3) BETWEEN "+filtroNumero1_var.get()+" and "+filtroNumero2_var.get()+" ORDER BY substring(`numero`, 5) ASC, substring(`numero`, 1, 3) ASC")

            elif filtroTipo_var.get()!='Todos' and filtroA_var.get()=='Todos':
                mycursor.execute("SELECT * FROM Transaccion WHERE `tipo` = '"+filtroTipo_var.get()+"' AND substring(`numero`, 1, 3) BETWEEN "+filtroNumero1_var.get()+" and "+filtroNumero2_var.get()+" ORDER BY substring(`numero`, 5) ASC, substring(`numero`, 1, 3) ASC")
            
            elif filtroTipo_var.get()=='Todos' and filtroA_var.get()!='Todos':
                mycursor.execute("SELECT * FROM Transaccion WHERE YEAR(fecha) = '"+filtroA_var.get()+"' AND substring(`numero`, 1, 3) BETWEEN "+filtroNumero1_var.get()+" and "+filtroNumero2_var.get()+" ORDER BY substring(`numero`, 5) ASC, substring(`numero`, 1, 3) ASC")

            else:
                mycursor.execute("SELECT * FROM Transaccion WHERE substring(`numero`, 1, 3) BETWEEN "+filtroNumero1_var.get()+" and "+filtroNumero2_var.get()+" ORDER BY substring(`numero`, 5) ASC, substring(`numero`, 1, 3) ASC, `tipo` ASC")

        else:
            if filtroTipo_var.get()!='Todos' and filtroA_var.get()!='Todos':
                mycursor.execute("SELECT * FROM Transaccion WHERE LOCATE('"+busqueda_var.get()+"', `asunto`) > 0 AND `tipo` = '"+filtroTipo_var.get()+"' AND YEAR(fecha) = '"+filtroA_var.get()+"' AND substring(`numero`, 1, 3) BETWEEN "+filtroNumero1_var.get()+" and "+filtroNumero2_var.get()+" ORDER BY substring(`numero`, 5) ASC, substring(`numero`, 1, 3) ASC")

            elif filtroTipo_var.get()!='Todos' and filtroA_var.get()=='Todos':
                mycursor.execute("SELECT * FROM Transaccion WHERE LOCATE('"+busqueda_var.get()+"', `asunto`) > 0 AND `tipo` = '"+filtroTipo_var.get()+"' AND substring(`numero`, 1, 3) BETWEEN "+filtroNumero1_var.get()+" and "+filtroNumero2_var.get()+" ORDER BY substring(`numero`, 5) ASC, substring(`numero`, 1, 3) ASC")
            
            elif filtroTipo_var.get()=='Todos' and filtroA_var.get()!='Todos':
                mycursor.execute("SELECT * FROM Transaccion WHERE LOCATE('"+busqueda_var.get()+"', `asunto`) > 0 AND YEAR(fecha) = '"+filtroA_var.get()+"' AND substring(`numero`, 1, 3) BETWEEN "+filtroNumero1_var.get()+" and "+filtroNumero2_var.get()+" ORDER BY substring(`numero`, 5) ASC, substring(`numero`, 1, 3) ASC")
            else:
                mycursor.execute("SELECT * FROM Transaccion WHERE LOCATE('"+busqueda_var.get()+"', `asunto`) > 0 AND substring(`numero`, 1, 3) BETWEEN "+filtroNumero1_var.get()+" and "+filtroNumero2_var.get()+" ORDER BY substring(`numero`, 5) ASC, substring(`numero`, 1, 3) ASC, `tipo` ASC")
        
        fila = mycursor.fetchall()
        mycursor.close()
        conexion_bdd.close()
        
        # Se insertan en la tabla los datos seleccionados
        for dato in fila:
            if dato['nCheque']!=0:
                tabla.insert('', 'end', values=(dato['numero'], dato['tipo'], dato['asunto'], dato['persona'], dato['fecha'].strftime("%d-%m-%Y"), dato['medio'], dato['nCheque'], '{:,}'.format(dato['monto']).replace(',','.'), dato['descripcion']))
            else: tabla.insert('', 'end', values=(dato['numero'], dato['tipo'], dato['asunto'], dato['persona'], dato['fecha'].strftime("%d-%m-%Y"), dato['medio'], "--------", '{:,}'.format(dato['monto']).replace(',','.'), dato['descripcion']))

        if filtroTipo_var.get()=='Ingreso':
            tabla.heading('4', text="Recibido de", anchor=W)
        elif filtroTipo_var.get()=='Egreso':
            tabla.heading('4', text="Enviado a", anchor=W)
        else: tabla.heading('4', text="Recibido de/Enviado a", anchor=W)

    except pymysql.Error as error:
        messagebox.showerror(title="Error", message=error)


# Actualización de datos de la base de datos
def actualizar_dato_baseDeDatos():
    try:
        conexion_bdd = pymysql.connect(user=DB_USER, password=DB_USER_PASSWORD, host=DB_HOST, database=DB_NAME, cursorclass=pymysql.cursors.DictCursor, ssl={"rejectUnauthorized":False})
        sql = "UPDATE Transaccion SET `asunto` = %s, `persona` = %s, `fecha` = %s, `medio` = %s, `nCheque` = %s, `monto` = %s, `descripcion` = %s WHERE `numero` = %s AND `tipo` = %s"
        valores = (asunto, persona, date.isoformat(fecha), medio, ncheque, monto, descripcion, numero, tipo)

        mycursor = conexion_bdd.cursor()
        mycursor.execute(sql, valores)
        conexion_bdd.commit()

        mycursor.close()
        conexion_bdd.close()


        if medio=='Cheque':
            tabla.item(elemento, values=(numero, tipo, asunto, persona, fecha.strftime("%d-%m-%Y"), medio, ncheque, '{:,}'.format(monto).replace(',','.'), descripcion))
        else: tabla.item(elemento, values=(numero, tipo, asunto, persona, fecha.strftime("%d-%m-%Y"), medio, "--------", '{:,}'.format(monto).replace(',','.'), descripcion))

    except pymysql.Error as error:
        if messagebox.showerror(title="Error", message=error):
            cerrar_seccion_editar()


# Exportación de datos de un mes-año específico a planilla excel
def exportar_datos_baseDeDatos():
    try:
        conexion_bdd = pymysql.connect(user=DB_USER, password=DB_USER_PASSWORD, host=DB_HOST, database=DB_NAME, cursorclass=pymysql.cursors.DictCursor, ssl={"rejectUnauthorized":False})
        mycursor = conexion_bdd.cursor()
        mycursor.execute("SELECT * FROM Transaccion WHERE MONTH(fecha) = "+mes[0:2]+" AND YEAR(fecha) = "+anio+" AND `tipo` = '"+tipo+"'")
        fila = mycursor.fetchall()
        mycursor.close()
        conexion_bdd.close()

        lista_numero=[]
        lista_asunto=[]
        lista_persona=[]
        lista_fecha=[]
        lista_medio=[]
        lista_ncheque=[]
        lista_monto=[]
        lista_descripcion=[]
        for dato in fila:
            lista_numero.append(dato['numero'])
            lista_asunto.append(dato['asunto'])
            lista_persona.append(dato['persona'])
            lista_fecha.append(dato['fecha'].strftime("%d-%m-%Y"))
            lista_medio.append(dato['medio'])
            if dato['nCheque']!=0:
                lista_ncheque.append(dato['nCheque'])
            else: lista_ncheque.append(" ")
            lista_monto.append(dato['monto'])
            lista_descripcion.append(dato['descripcion'])
        
        if tipo=='Ingreso':
            data = pd.DataFrame({'Numero':lista_numero, 'Asunto':lista_asunto, 'Recibido de':lista_persona, 'Fecha':lista_fecha, 'Medio':lista_medio, 'N° Cheque':lista_ncheque, 'Monto':lista_monto, 'Por concepto de':lista_descripcion})
        else: 
            data = pd.DataFrame({'Numero':lista_numero, 'Asunto':lista_asunto, 'Para':lista_persona, 'Fecha':lista_fecha, 'Medio':lista_medio, 'N° Cheque':lista_ncheque, 'Monto':lista_monto, 'Por concepto de':lista_descripcion})

        filepath = findfile(tipo+"s_"+anio+"_PLANTILLA.xlsx", "\\")
        if filepath==None:
            libro=xls.Workbook(tipo+"s_"+anio+"_PLANTILLA.xlsx")
            libro.close()
            with pd.ExcelWriter(tipo+"s_"+anio+"_PLANTILLA.xlsx") as writer:
                data.to_excel(writer, sheet_name=mes[5:])
        else:
            rod = os.path.dirname(os.path.abspath(filepath))
            wb = load_workbook(rod+"\\"+tipo+"s_"+anio+"_PLANTILLA.xlsx", read_only=True)

            if not mes[5:] in wb.sheetnames:
                try:
                    with pd.ExcelWriter(rod+"\\"+tipo+"s_"+anio+"_PLANTILLA.xlsx", mode="a", engine="openpyxl") as writer:
                        data.to_excel(writer, sheet_name=mes[5:])
                except:
                    messagebox.showerror(title="Error", message="Cierre el archivo excel "+tipo+"s_"+anio+"_PLANTILLA.xlsx antes de exportar el mes de "+mes[5:])

            wb.close()

    except pymysql.Error as error:
        if messagebox.showerror(title="Error", message=error):
            cerrar_seccion_exportar()


#=================================== FUNCIONES DE AGREGAR INGRESO/EGRESO ======================================#
# Inicialización de variables
def inicializar_variables():
    global numero_var
    global asunto_var
    global asuntoOtro_var
    global persona_var
    global personaOtra_var 
    global medio_var
    global ncheque_var
    global monto_var
    global descripcion_var
    global imprimir
    numero_var=StringVar()
    asunto_var=StringVar()
    asuntoOtro_var=StringVar()
    persona_var=StringVar()
    personaOtra_var=StringVar() 
    medio_var=StringVar()
    ncheque_var=StringVar()
    monto_var=StringVar()
    descripcion_var=StringVar()
    imprimir=BooleanVar()

# Inicialización de componentes
def inicializar_componentes(tipo):
    global contenedor0
    global contenedor1
    global contenedor2
    global contenedor3
    global contenedor4
    global contenedor5
    global contenedor6

    global entrada0
    global entrada1
    global entrada1_adicional
    global entrada2
    global entrada2_adicional
    global entrada3
    global entrada4
    global entrada5
    global entrada6

    global botonCancelar
    global botonGuardar

    # Validación entradas vacias
    def validacion_vacia(evento):
        if medio_var.get()=='Cheque':
            if entrada1.get()=='OTRO' and entrada2.get()=='OTRO':
                if len(entrada1_adicional.get())>0 and len(entrada2_adicional.get())>0 and len(entrada4.get())>0 and len(entrada5.get())>0 and len(entrada6.get("1.0", "end-1c")):
                    botonGuardar['state']=NORMAL
                else: botonGuardar['state']=DISABLED
            elif entrada1.get()=='OTRO' and entrada2.get()!='OTRO':
                if len(entrada1_adicional.get())>0 and len(entrada2.get())>0 and len(entrada4.get())>0 and len(entrada5.get())>0 and len(entrada6.get("1.0", "end-1c")):
                    botonGuardar['state']=NORMAL
                else: botonGuardar['state']=DISABLED
            elif entrada1.get()!='OTRO' and entrada2.get()=='OTRO':
                if len(entrada1.get())>0 and len(entrada2_adicional.get())>0 and len(entrada4.get())>0 and len(entrada5.get())>0 and len(entrada6.get("1.0", "end-1c")):
                    botonGuardar['state']=NORMAL
                else: botonGuardar['state']=DISABLED
            else:
                if len(entrada1.get())>0 and len(entrada2.get())>0 and len(entrada4.get())>0 and len(entrada5.get())>0 and len(entrada6.get("1.0", "end-1c")):
                    botonGuardar['state']=NORMAL
                else: botonGuardar['state']=DISABLED
        else:
            if entrada1.get()=='OTRO' and entrada2.get()=='OTRO':
                if len(entrada1_adicional.get())>0 and len(entrada2_adicional.get())>0 and len(entrada5.get())>0 and len(entrada6.get("1.0", "end-1c")):
                    botonGuardar['state']=NORMAL
                else: botonGuardar['state']=DISABLED
            elif entrada1.get()=='OTRO' and entrada2.get()!='OTRO':
                if len(entrada1_adicional.get())>0 and len(entrada2.get())>0 and len(entrada5.get())>0 and len(entrada6.get("1.0", "end-1c")):
                    botonGuardar['state']=NORMAL
                else: botonGuardar['state']=DISABLED
            elif entrada1.get()!='OTRO' and entrada2.get()=='OTRO':
                if len(entrada1.get())>0 and len(entrada2_adicional.get())>0 and len(entrada5.get())>0 and len(entrada6.get("1.0", "end-1c")):
                    botonGuardar['state']=NORMAL
                else: botonGuardar['state']=DISABLED
            else:
                if len(entrada1.get())>0 and len(entrada2.get())>0 and len(entrada5.get())>0 and len(entrada6.get("1.0", "end-1c")):
                    botonGuardar['state']=NORMAL
                else: botonGuardar['state']=DISABLED

    app.bind('<KeyRelease>', validacion_vacia)


    # Label Frame
    contenedor0=LabelFrame(contenedor_campos, text="Número de folio", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedor0.place(x=10, y=10, width=230, height=65)

    contenedor3=LabelFrame(contenedor_campos, text="Fecha", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedor3.place(x=250, y=10, width=340, height=65)

    contenedor1=LabelFrame(contenedor_campos, text="Asunto", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedor1.place(x=10, y=80, width=580, height=105)

    if tipo=='Ingreso':
        contenedor2=LabelFrame(contenedor_campos, text="Recibido de", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    else: contenedor2=LabelFrame(contenedor_campos, text="Enviado a", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedor2.place(x=10, y=190, width=580, height=105)

    contenedor4=LabelFrame(contenedor_campos, text="Número de Cheque", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedor4.place(x=10, y=370, width=285, height=65)

    contenedor5=LabelFrame(contenedor_campos, text="Monto", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedor5.place(x=305, y=370, width=285, height=65)

    contenedor6=LabelFrame(contenedor_campos, text="Por concepto de", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedor6.place(x=10, y=440, width=580, height=150)

    # Entradas
    def obtener_numero():
        global fecha
        global tipoT
        fecha=entrada3.get_date()
        tipoT=tipo
        obtener_numeroDeFolio_baseDeDatos()
        numero_var.set(numero)
    def dateentryclick(evento):
        obtener_numero()
    entrada3=dateentry.DateEntry(contenedor3, font=("Helvetica", 13), state='readonly', locale='es_CL', date_pattern='dd-mm-yyyy', width=50)
    entrada3.set_date(fecha_anterior)
    entrada3.bind("<<DateEntrySelected>>", dateentryclick)
    entrada3.place(x=10, y=5, width=320, height=32)
    entrada3.config(headersbackground="#E62B0A", headersforeground="#ffffff", foreground="#000000", background="#ffffff")
    entrada0=Entry(contenedor0, textvariable=numero_var, font=("Helvetica", 13), state='readonly')
    entrada0.place(x=10, y=5, width=210, height=32)
    obtener_numero()


    def comboboxclick1(evento):
        if asunto_var.get() !='OTRO':
            entrada1_adicional.place_forget()
        else:
            entrada1_adicional.place(x=10, y=45, width=560, height=32)
    entrada1=ttk.Combobox(contenedor1, textvariable=asunto_var, font=("Helvetica", 13), state='readonly')
    if tipo=="Ingreso":
        entrada1['values']=lista_asunto_ingreso
    else:
        entrada1['values']=lista_asunto_egreso
    entrada1.bind("<<ComboboxSelected>>", comboboxclick1)
    entrada1.place(x=10, y=5, width=560, height=32)
    entrada1_adicional=Entry(contenedor1, textvariable=asuntoOtro_var, font=("Helvetica", 13))

    def comboboxclick2(evento):
        if persona_var.get() !='OTRO':
            entrada2_adicional.place_forget()
        else:
            entrada2_adicional.place(x=10, y=45, width=560, height=32)

    entrada2=ttk.Combobox(contenedor2, textvariable=persona_var, font=("Helvetica", 13), state='readonly')
    if tipo=="Ingreso":
        entrada2['values']=lista_recibido_de
    else:
        entrada2['values']=lista_enviado_a
    entrada2.bind("<<ComboboxSelected>>", comboboxclick2)
    entrada2.place(x=10, y=5, width=560, height=32)
    entrada2_adicional=Entry(contenedor2, textvariable=personaOtra_var, font=("Helvetica", 13))

    # Validación entradas solo números
    def validacion_numeros(entrada):
        return entrada.isdigit()

    validacionNumero=contenedor_campos.register(validacion_numeros)

    entrada4=Entry(contenedor4, textvariable=ncheque_var, font=("Helvetica", 13), validate="key", validatecommand=(validacionNumero, '%S'))
    entrada4.place(x=10, y=5, width=265, height=32)

    # Mostrar formato moneda miles
    def mostrar_formato(*args):
        if len(entrada5.get())>0:
            monto_var.set('{:,}'.format(int(entrada5.get())).replace(",","."))

    # Quitar formato moneda miles
    def quitar_formato(*args):
        if len(entrada5.get())>0:
            monto_var.set(entrada5.get().replace(".",""))
            entrada5['validate']="key"
            entrada5['validatecommand']=(validacionNumero, '%S')

    entrada5=Entry(contenedor5, textvariable=monto_var, font=("Helvetica", 13), validate="key", validatecommand=(validacionNumero, '%S'))
    entrada5.place(x=10, y=5, width=265, height=32)
    entrada5.bind("<FocusOut>", mostrar_formato)
    entrada5.bind("<FocusIn>", quitar_formato)


    barra1=Scrollbar(contenedor6)
    barra1.place(x=555, y=5, height=110)
    entrada6=Text(contenedor6, wrap=WORD, font=("Helvetica", 13), yscrollcommand = barra1.set)
    entrada6.place(x=10, y=5, width=540, height=110)
    barra1.config(command=entrada6.yview)

    # Label Frame
    contenedorRB=LabelFrame(contenedor_campos, text="Medio", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedorRB.place(x=10, y=300, width=580, height=65)

    def mostrar_contenedor4():
        if medio_var.get()=='Cheque':
            contenedor4.place(x=10, y=370, width=285, height=65)
        else: contenedor4.place_forget()

    # Botones radio
    radioBoton1=Radiobutton(contenedorRB, text="Cheque", font=("Helvetica", 13), variable=medio_var, value="Cheque", command=mostrar_contenedor4, bg='#FFFFFF')
    radioBoton1.place(x=10, y=5)
    radioBoton1.select()

    radioBoton2=Radiobutton(contenedorRB, text="Efectivo", font=("Helvetica", 13), variable=medio_var, value="Efectivo", command=mostrar_contenedor4, bg='#FFFFFF')
    radioBoton2.place(x=185, y=5)
    
    radioBoton3=Radiobutton(contenedorRB, text="Transferencia", font=("Helvetica", 13), variable=medio_var, value="Transferencia", command=mostrar_contenedor4, bg='#FFFFFF')
    radioBoton3.place(x=360, y=5)


    # CheckBox
    checkBox=Checkbutton(contenedor_campos, text="¿Desea imprimir los datos del "+tipo+"?", variable=imprimir, onvalue=TRUE, offvalue=FALSE, font=("Helvetica", 12))
    checkBox.configure(bg='#FFFFFF')
    checkBox.place(x=10, y=610)

    # Botones
    botonCancelar=Button(contenedor_campos, text="Cancelar", command=lambda:cerrar_seccion_agregar(), font=("Helvetica", 13), borderwidth=3)
    botonCancelar.place(x=410, y=610)
    botonGuardar=Button(contenedor_campos, text="Guardar", command=lambda:crearTransaccion(tipo), font=("Helvetica", 13), borderwidth=3)
    botonGuardar.place(x=510, y=610)
    botonGuardar['state']=DISABLED

# Cierre sección agregar
def cerrar_seccion_agregar():
    entradaBuscar['state']=NORMAL
    botonLimpiar['state']=NORMAL
    combo_tipo['state']='readonly'
    combo_anio['state']='readonly'
    combo_numero1['state']='readonly'
    combo_numero2['state']='readonly'
    contenedor4.place_forget()
    contenedor_campos.place_forget()
    contenedor_operaciones.place(x=890, y=10, width=600, height=75)

# Creación del nuevo elemento
def crearTransaccion(t):
    global tipo
    global asunto
    global persona
    global fecha
    global medio
    global ncheque
    global monto
    global descripcion
    asunto=entrada1.get()
    if asunto=='OTRO':
        asunto=entrada1_adicional.get()
    persona=entrada2.get()
    if persona=='OTRO':
        persona=entrada2_adicional.get()
    fecha=entrada3.get_date()
    medio=medio_var.get()
    if medio=="Cheque":
        ncheque=int(entrada4.get())
    else: ncheque=0
    monto=int(entrada5.get().replace('.',''))
    descripcion=entrada6.get("1.0", "end-1c")
    tipo=t
    
    insertar_dato_baseDeDatos() # Conexión a la base de datos (agregar elemento)
    cerrar_seccion_agregar()
    crear_documento()
    if imprimir.get()==True:
        imprimir_documento()
    global fecha_anterior
    fecha_anterior=entrada3.get() # Se guarda la fecha anterior

#======================================== FUNCIONES DE EDITAR INGRESO/EGRESO ==========================================================#
# Inicialización de variables edición
def inicializar_variables_editor():
    global numero_var
    global asunto_var
    global persona_var
    global fecha_var
    global medio_var
    global ncheque_var
    global monto_var
    global descripcion_var
    global imprimir
    numero_var=StringVar()
    asunto_var=StringVar()
    persona_var=StringVar()
    fecha_var=StringVar()
    medio_var=StringVar()
    ncheque_var=StringVar()
    monto_var=StringVar()
    descripcion_var=StringVar()
    imprimir=BooleanVar()

# Inicialización componentes editor
def inicializar_componentes_editor(tipo):
    global contenedor0
    global contenedor1
    global contenedor2
    global contenedor3
    global contenedor4
    global contenedor5
    global contenedor6

    global entrada0
    global entrada1
    global entrada1_adicional
    global entrada2
    global entrada2_adicional
    global entrada3
    global entrada4
    global entrada5
    global entrada6

    global contenedorMedio

    global botonCancelar
    global botonGuardar

    # Botones
    botonCancelar=Button(contenedor_editor, text="Cancelar", command=cerrar_seccion_editar, font=("Helvetica", 13), borderwidth=3)
    botonCancelar.place(x=410, y=550)
    botonGuardar=Button(contenedor_editor, text="Guardar", command=guardar_cambios_edicion, font=("Helvetica", 13), borderwidth=3)
    botonGuardar.place(x=510, y=550)
    botonGuardar['state']=DISABLED

    medio_var.set(tabla.item(elemento)['values'][5])

    def validacion_datos(evento):

        if medio_var.get()=='Cheque':
            if (len(entrada1.get())>0 and len(entrada2.get())>0 and len(entrada4.get())>0 and len(entrada5.get())>0 and len(entrada6.get("1.0", "end-1c"))>0) and (entrada1.get()!=tabla.item(elemento)['values'][2] or entrada2.get()!=tabla.item(elemento)['values'][3] or entrada4.get()!=str(tabla.item(elemento)['values'][6]) or entrada5.get().replace('.','')!=str(tabla.item(elemento)['values'][7]).replace('.','') or entrada6.get("1.0", "end-1c")!=tabla.item(elemento)['values'][8] or entrada3.get_date().strftime("%d-%m-%Y")!=tabla.item(elemento)['values'][4] or medio_var.get()!=tabla.item(elemento)['values'][5]) and entrada3.get_date().strftime("%Y")==tabla.item(elemento)['values'][4][-4:]:
                botonGuardar['state']=NORMAL
            else: botonGuardar['state']=DISABLED
        else:
            if len(entrada1.get())>0 and len(entrada2.get())>0 and len(entrada5.get())>0 and len(entrada6.get("1.0", "end-1c"))>0 and (entrada1.get()!=tabla.item(elemento)['values'][2] or entrada2.get()!=tabla.item(elemento)['values'][3] or entrada5.get().replace('.','')!=str(tabla.item(elemento)['values'][7]).replace('.','') or entrada6.get("1.0", "end-1c")!=tabla.item(elemento)['values'][8] or entrada3.get_date().strftime("%d-%m-%Y")!=tabla.item(elemento)['values'][4] or medio_var.get()!=tabla.item(elemento)['values'][5]) and entrada3.get_date().strftime("%Y")==tabla.item(elemento)['values'][4][-4:]:
                botonGuardar['state']=NORMAL
            else: botonGuardar['state']=DISABLED

    app.bind('<KeyRelease>', validacion_datos)

    # Label Frame
    contenedor0=LabelFrame(contenedor_editor, text="Número de folio", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedor0.place(x=10, y=10, width=230, height=65)

    contenedor3=LabelFrame(contenedor_editor, text="Fecha", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedor3.place(x=250, y=10, width=340, height=65)

    contenedor1=LabelFrame(contenedor_editor, text="Asunto", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedor1.place(x=10, y=80, width=580, height=65)

    if tipo=='Ingreso':
        contenedor2=LabelFrame(contenedor_editor, text="Recibido de", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    else: contenedor2=LabelFrame(contenedor_editor, text="Enviado a", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedor2.place(x=10, y=150, width=580, height=65)

    contenedor4=LabelFrame(contenedor_editor, text="Número de Cheque", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')

    contenedor5=LabelFrame(contenedor_editor, text="Monto", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedor5.place(x=305, y=290, width=285, height=65)

    contenedor6=LabelFrame(contenedor_editor, text="Por concepto de", font=("Helvetica", 13), bg='#D4CDD6', fg='#02020D')
    contenedor6.place(x=10, y=360, width=580, height=150)


    entrada3=dateentry.DateEntry(contenedor3, font=("Helvetica", 13), state='readonly', locale='es_CL', date_pattern='dd-mm-yyyy', width=50)
    entrada3.bind("<<DateEntrySelected>>", validacion_datos)
    entrada3.place(x=10, y=5, width=320, height=32)
    entrada3.set_date(tabla.item(elemento)['values'][4])
    

    entrada0=Entry(contenedor0, textvariable=numero_var, font=("Helvetica", 13), state='readonly')
    entrada0.place(x=10, y=5, width=210, height=32)
    numero_var.set(tabla.item(elemento)['values'][0])

    entrada1=Entry(contenedor1, textvariable=asunto_var, font=("Helvetica", 13))
    entrada1.place(x=10, y=5, width=560, height=32)
    entrada1.insert(0, tabla.item(elemento)['values'][2])

    entrada2=Entry(contenedor2, textvariable=persona_var, font=("Helvetica", 13))
    entrada2.place(x=10, y=5, width=560, height=32)
    entrada2.insert(0, tabla.item(elemento)['values'][3])

    # Validación entradas solo números
    def validacion_numeros(entrada):
        return entrada.isdigit()

    validacionNumero=contenedor_campos.register(validacion_numeros)

    entrada4=Entry(contenedor4, textvariable=ncheque_var, font=("Helvetica", 13), validate="key", validatecommand=(validacionNumero, '%S'))
    entrada4.place(x=10, y=5, width=265, height=32)

    def mostrar_formato(*args):
        if len(entrada5.get())>0:
            monto_var.set('{:,}'.format(int(entrada5.get())).replace(",","."))

    def quitar_formato(*args):
        if len(entrada5.get())>0:
            monto_var.set(entrada5.get().replace(".",""))
            entrada5['validate']="key"
            entrada5['validatecommand']=(validacionNumero, '%S')
    
    entrada5=Entry(contenedor5, textvariable=monto_var, font=("Helvetica", 13))
    entrada5.place(x=10, y=5, width=265, height=32)
    entrada5.insert(0, tabla.item(elemento)['values'][7])
    entrada5.bind("<FocusOut>", mostrar_formato)
    entrada5.bind("<FocusIn>", quitar_formato)


    barra1=Scrollbar(contenedor6)
    barra1.place(x=555, y=5, height=110)
    entrada6=Text(contenedor6, wrap=WORD, font=("Helvetica", 13), yscrollcommand = barra1.set)
    entrada6.place(x=10, y=5, width=540, height=110)
    entrada6.insert("1.0", tabla.item(elemento)['values'][8])
    barra1.config(command=entrada6.yview)

    # Medio
    contenedorMedio=LabelFrame(contenedor_editor, text="Medio", font=("Helvetica", 12), bg='#D4CDD6', fg='#02020D')
    contenedorMedio.place(x=10, y=220, width=580, height=65)  

    def mostrar_contenedor4():
        if medio_var.get()=='Cheque':
            contenedor4.place(x=10, y=290, width=285, height=65)
            if len(entrada4.get())>0:
                entrada4.delete(0, 'end')
            entrada4.insert(0, str(tabla.item(elemento)['values'][6]))
        else:
            entrada4.delete(0, 'end')
            contenedor4.place_forget()
        validacion_datos(None)
    
    # Botones radio
    radioBoton1=Radiobutton(contenedorMedio, text="Cheque", font=("Helvetica", 13), variable=medio_var, value="Cheque", command=mostrar_contenedor4, bg='#FFFFFF')
    radioBoton1.place(x=10, y=5)

    radioBoton2=Radiobutton(contenedorMedio, text="Efectivo", font=("Helvetica", 13), variable=medio_var, value="Efectivo", command=mostrar_contenedor4, bg='#FFFFFF')
    radioBoton2.place(x=185, y=5)
    
    radioBoton3=Radiobutton(contenedorMedio, text="Transferencia", font=("Helvetica", 13), variable=medio_var, value="Transferencia", command=mostrar_contenedor4, bg='#FFFFFF')
    radioBoton3.place(x=360, y=5)

    if tabla.item(elemento)['values'][5]=='Cheque':
        mostrar_contenedor4()
        radioBoton1.select()
    elif tabla.item(elemento)['values'][5]=='Efectivo':
        radioBoton2.select()
    else:
        radioBoton3.select()


    # CheckBox
    checkBox=Checkbutton(contenedor_editor, text="¿Desea imprimir los datos del "+tipo+"?", variable=imprimir, onvalue=TRUE, offvalue=FALSE, font=("Helvetica", 12))
    checkBox.configure(bg='#FFFFFF')
    checkBox.place(x=10, y=550)

# Cierre sección editar
def cerrar_seccion_editar():
    entradaBuscar['state']=NORMAL
    botonLimpiar['state']=NORMAL
    combo_tipo['state']='readonly'
    combo_anio['state']='readonly'
    combo_numero1['state']='readonly'
    combo_numero2['state']='readonly'
    contenedor4.place_forget()
    contenedor_editor.place_forget()
    contenedor_operaciones.place(x=890, y=10, width=600, height=75)
    tabla.selection_remove(tabla.selection())
    botonAgregarIngreso.place(x=20, y=10)
    botonAgregarEgreso.place(x=180, y=10)
    botonExportarExcel.place(x=340, y=10)
    botonEditar.place_forget()
    botonImprimir.place_forget()

# Se Guardan los cambios realizados
def guardar_cambios_edicion():
    global numero
    global tipo
    global asunto
    global persona
    global fecha
    global medio
    global ncheque
    global monto
    global descripcion
    asunto=entrada1.get()
    persona=entrada2.get()
    fecha=entrada3.get_date()
    medio=medio_var.get()
    if medio=="Cheque":
        ncheque=int(entrada4.get())
    else: ncheque=0
    monto=int(entrada5.get().replace('.',''))
    descripcion=entrada6.get("1.0", "end-1c")
    tipo=tabla.item(elemento)['values'][1]
    numero=tabla.item(elemento)['values'][0]

    
    actualizar_dato_baseDeDatos() # Conexión a la base de datos (editar elemento)
    cerrar_seccion_editar()
    crear_documento()
    if imprimir.get()==True:
        imprimir_documento()

# Función que busca la ruta del archivo 
def findfile(name, path):
    for dirpath, dirname, filename in os.walk(path):
        if name in filename:
            return os.path.join(dirpath, name)
    
# Función que crea el documento con los datos el elemento nuevo a partir de una plantilla
def crear_documento():
    if tipo=='Ingreso':
        #filepath = findfile("plantilla_documento_ingreso.docx", "\\")
        plantilla_documento = "plantilla_documento_ingreso.docx"
    else:
        #filepath = findfile("plantilla_documento_egreso.docx", "\\")
        plantilla_documento = "plantilla_documento_egreso.docx"
    
    documento = DocxTemplate(plantilla_documento)

    f=Formato()
    monto_en_palabras=f.numero_a_moneda_sunat(monto) # monto en palabras

    # Datos a insertar en el documento
    if medio=='Cheque':
        contexto = {
            "NUMERO": numero,
            "ASUNTO": asunto,
            "PERSONA": persona,
            "FECHA": fecha.strftime("%d-%m-%Y"),
            "CHEQUE": ncheque,
            "E": "",
            "T": "",
            "MONTO": '{:,}'.format(monto).replace(',','.'),
            "MONTO_PALABRAS": monto_en_palabras,
            "DESCRIPCION": descripcion,
        }
    elif medio=='Efectivo':
        contexto = {
            "NUMERO": numero,
            "ASUNTO": asunto,
            "PERSONA": persona,
            "FECHA": fecha.strftime("%d-%m-%Y"),
            "CHEQUE": "",
            "E": "X",
            "T": "",
            "MONTO": '{:,}'.format(monto).replace(',','.'),
            "MONTO_PALABRAS": monto_en_palabras,
            "DESCRIPCION": descripcion,
        }
    else:
        contexto = {
            "NUMERO": numero,
            "ASUNTO": asunto,
            "PERSONA": persona,
            "FECHA": fecha.strftime("%d-%m-%Y"),
            "CHEQUE": "",
            "E": "",
            "T": "X",
            "MONTO": '{:,}'.format(monto).replace(',','.'),
            "MONTO_PALABRAS": monto_en_palabras,
            "DESCRIPCION": descripcion,
        }
    
    documento.render_init()
    documento.render(contexto, autoescape=True)
    docx_filename=numero+"_"+tipo+".docx"
    documento.save(docx_filename)
    pdf_filename=docx_filename.replace(".docx", ".pdf")
    convert(docx_filename, pdf_filename)
    

# Función que Imprime el documento en la impresora predeterminada
def imprimir_documento():
    name = win32print.GetDefaultPrinter()
    printdefaults = {"DesiredAccess": win32print.PRINTER_ALL_ACCESS}
    handle = win32print.OpenPrinter(name, printdefaults)
    level = 2
    attributes = win32print.GetPrinter(handle, level)
    attributes['pDevMode'].Duplex = 3 
    win32print.SetPrinter(handle, level, attributes, 0)
    win32print.GetPrinter(handle, level)['pDevMode'].Duplex
    win32api.ShellExecute(0, "print", str(numero)+"_"+tipo+".pdf", None,  ".",  0)
    win32print.ClosePrinter(handle)

#=================================== FUNCIONES DE EXPORTAR A EXCEL =========================================#
# Inicialización de variables exportar
def inicializar_variables_exportar():
    global tipo_var
    global mes_var
    global a_var
    tipo_var=StringVar()
    mes_var=StringVar()
    a_var=StringVar()

# Inicialización de componentes
def inicializar_componentes_exportar():
    
    contenedorTipo=LabelFrame(contenedor_exportar, text="Tipo", font=("Helvetica", 12), bg='#D4CDD6', fg='#02020D')
    contenedorTipo.place(x=10, y=10, width=240, height=65)

    contenedorMes=LabelFrame(contenedor_exportar, text="Mes", font=("Helvetica", 12), bg='#D4CDD6', fg='#02020D')
    contenedorMes.place(x=260, y=10, width=220, height=65)

    contenedorA=LabelFrame(contenedor_exportar, text="Año", font=("Helvetica", 12), bg='#D4CDD6', fg='#02020D')
    contenedorA.place(x=490, y=10, width=100, height=65)

    radioBotonIngresos=Radiobutton(contenedorTipo, text="Ingresos", font=("Helvetica", 13), variable=tipo_var, value="Ingreso", bg='#FFFFFF')
    radioBotonIngresos.place(x=10, y=5)
    radioBotonIngresos.select()

    radioBotonEgresos=Radiobutton(contenedorTipo, text="Egresos", font=("Helvetica", 13), variable=tipo_var, value="Egreso", bg='#FFFFFF')
    radioBotonEgresos.place(x=130, y=5)

    def validacion_entradas(evento):
        if len(entradaMes.get())>0 and len(entradaA.get())>0:
            botonExportar['state']=NORMAL
        else: botonExportar['state']=DISABLED

    entradaMes=ttk.Combobox(contenedorMes, textvariable=mes_var, font=("Helvetica", 13), state='readonly', values=lista_meses)
    entradaMes.place(x=10, y=5, width=200, height=32)
    entradaMes.bind("<<ComboboxSelected>>", validacion_entradas)

    entradaA=ttk.Combobox(contenedorA, textvariable=a_var, font=("Helvetica", 13), state='readonly', values=lista_a)
    entradaA.place(x=10, y=5, width=80, height=32)
    entradaA.bind("<<ComboboxSelected>>", validacion_entradas)

    # Botones
    botonCancelar=Button(contenedor_exportar, text="Cancelar", command=cerrar_seccion_exportar, font=("Helvetica", 13), borderwidth=3)
    botonCancelar.place(x=410, y=100)
    botonExportar=Button(contenedor_exportar, text="Exportar", command=exportar_datos, font=("Helvetica", 13), borderwidth=3)
    botonExportar.place(x=510, y=100)
    botonExportar['state']=DISABLED

# Cierre sección
def cerrar_seccion_exportar():
    entradaBuscar['state']=NORMAL
    botonLimpiar['state']=NORMAL
    combo_tipo['state']='readonly'
    combo_anio['state']='readonly'
    combo_numero1['state']='readonly'
    combo_numero2['state']='readonly'
    contenedor_exportar.place_forget()
    contenedor_operaciones.place(x=890, y=10, width=600, height=75)

# Exportación de datos
def exportar_datos():
    cerrar_seccion_exportar()
    global tipo
    global mes
    global anio
    tipo=tipo_var.get()
    mes=mes_var.get()
    anio=a_var.get()
    exportar_datos_baseDeDatos()

# Cierre ventana principal
def cerrar_ventanaPrincipal():
    if messagebox.askokcancel("Salir", "¿Desea Salir?"):
        app.destroy()

#==================================== FUNCIONES DE CONFIGURACIÓN INICIAL =================================#
def agregar_ingreso():
    inicializar_variables()
    inicializar_componentes("Ingreso")
    entradaBuscar['state']=DISABLED
    botonBuscar['state']=DISABLED
    botonLimpiar['state']=DISABLED
    combo_tipo['state']=DISABLED
    combo_anio['state']=DISABLED
    combo_numero1['state']=DISABLED
    combo_numero2['state']=DISABLED
    contenedor_operaciones.place_forget()
    contenedor_campos['text']="Ingreso"
    contenedor_campos.place(x=890, y=10, width=600, height=680)

def agregar_egreso():
    inicializar_variables()
    inicializar_componentes("Egreso")
    entradaBuscar['state']=DISABLED
    botonBuscar['state']=DISABLED
    botonLimpiar['state']=DISABLED
    combo_tipo['state']=DISABLED
    combo_anio['state']=DISABLED
    combo_numero1['state']=DISABLED
    combo_numero2['state']=DISABLED
    contenedor_operaciones.place_forget()
    contenedor_campos['text']="Egreso"
    contenedor_campos.place(x=890, y=10, width=600, height=680)

def editar_transaccion():
    entradaBuscar['state']=DISABLED
    botonBuscar['state']=DISABLED
    botonLimpiar['state']=DISABLED
    combo_tipo['state']=DISABLED
    combo_anio['state']=DISABLED
    combo_numero1['state']=DISABLED
    combo_numero2['state']=DISABLED
    global elemento
    elemento=tabla.selection()[0]
    contenedor_operaciones.place_forget()
    contenedor_editor['text']="Editor "+tabla.item(elemento)['values'][1]
    contenedor_editor.place(x=890, y=10, width=600, height=680)
    inicializar_variables_editor()
    inicializar_componentes_editor(tabla.item(elemento)['values'][1])

def imprimir_transaccion():
    elemento=tabla.selection()[0]
    global numero
    global tipo
    global asunto
    global persona
    global fecha
    global medio
    global ncheque
    global monto
    global descripcion
    numero=tabla.item(elemento)['values'][0]
    tipo=tabla.item(elemento)['values'][1]
    asunto=tabla.item(elemento)['values'][2]
    persona=tabla.item(elemento)['values'][3]
    fecha=datetime.strptime(tabla.item(elemento)['values'][4], '%d-%m-%Y')
    medio=tabla.item(elemento)['values'][5]
    ncheque=tabla.item(elemento)['values'][6]
    monto=int(str(tabla.item(elemento)['values'][7]).replace(".",""))
    descripcion=tabla.item(elemento)['values'][8]
    #tabla.selection_remove(tabla.selection())
    deseleccionar_elemento(None)
    filepath=findfile(numero+"_"+tipo+".pdf", "\\")
    if filepath==None:
        crear_documento()
    imprimir_documento()
    

def exportar_a_excel():
    inicializar_variables_exportar()
    inicializar_componentes_exportar()
    entradaBuscar['state']=DISABLED
    botonBuscar['state']=DISABLED
    botonLimpiar['state']=DISABLED
    combo_tipo['state']=DISABLED
    combo_anio['state']=DISABLED
    combo_numero1['state']=DISABLED
    combo_numero2['state']=DISABLED
    contenedor_operaciones.place_forget()
    contenedor_exportar['text']="Exportación a Excel"
    contenedor_exportar.place(x=890, y=10, width=600, height=180)



def buscar_asunto():
    global fecha_anterior
    busqueda_var.set(busqueda_var.get().upper())

    if filtroA_var.get()!='Todos':
        fecha_anterior=fecha_anterior[0:len(fecha_anterior)-4]+filtroA_var.get()
        if filtroTipo_var.get()=='Ingreso':
            botonAgregarIngreso.place(x=20, y=10)
            botonAgregarEgreso.place_forget()
        elif filtroTipo_var.get()=='Egreso':
            botonAgregarIngreso.place_forget()
            botonAgregarEgreso.place(x=180, y=10)
        else:
            botonAgregarIngreso.place(x=20, y=10)
            botonAgregarEgreso.place(x=180, y=10)
    else:
        fecha_anterior=fecha_anterior[0:len(fecha_anterior)-4]+datetime.now().date().strftime('%Y')
        if filtroTipo_var.get()=='Ingreso':
            botonAgregarIngreso.place(x=20, y=10)
            botonAgregarEgreso.place_forget()
        elif filtroTipo_var.get()=='Egreso':
            botonAgregarIngreso.place_forget()
            botonAgregarEgreso.place(x=180, y=10)
        else:
            botonAgregarIngreso.place(x=20, y=10)
            botonAgregarEgreso.place(x=180, y=10)
    
    buscar_filtrar_baseDeDatos()

def limpiar_tabla():
    botonBuscar['state']=DISABLED
    busqueda_var.set('')
    filtroTipo_var.set('Todos')
    filtroA_var.set('Todos')
    filtroNumero1_var.set('000')
    filtroNumero2_var.set('000')
    contenedor_operaciones.place(x=890, y=10, width=600, height=75)
    limpiar_busqueda_baseDeDatos()

def filtrar_tabla(evento):
    global fecha_anterior
    busqueda_var.set(busqueda_var.get().upper())

    if filtroA_var.get()!='Todos':
        fecha_anterior=fecha_anterior[0:len(fecha_anterior)-4]+filtroA_var.get()
        if filtroTipo_var.get()=='Ingreso':
            botonAgregarIngreso.place(x=20, y=10)
            botonAgregarEgreso.place_forget()
        elif filtroTipo_var.get()=='Egreso':
            botonAgregarIngreso.place_forget()
            botonAgregarEgreso.place(x=180, y=10)
        else:
            botonAgregarIngreso.place(x=20, y=10)
            botonAgregarEgreso.place(x=180, y=10)
    else:
        fecha_anterior=fecha_anterior[0:len(fecha_anterior)-4]+datetime.now().date().strftime('%Y')
        if filtroTipo_var.get()=='Ingreso':
            botonAgregarIngreso.place(x=20, y=10)
            botonAgregarEgreso.place_forget()
        elif filtroTipo_var.get()=='Egreso':
            botonAgregarIngreso.place_forget()
            botonAgregarEgreso.place(x=180, y=10)
        else:
            botonAgregarIngreso.place(x=20, y=10)
            botonAgregarEgreso.place(x=180, y=10)

    if filtroNumero1_var.get()>filtroNumero2_var.get():
        if messagebox.showwarning(title="Advertencia", message="El rango inferior no puede ser mayor que el superior"):
            filtroNumero1_var.set('000')
            filtroNumero2_var.set('999')
    
    buscar_filtrar_baseDeDatos()


#====================================VENTANA PRINCIPAL=========================================
app=Tk()
app.title("Sistema de Caja")
width = 1500 # ancho de la ventana
height = 700 # alto de la ventana
x = 10
y = 10
app.geometry(f'{width}x{height}+{int(x)}+{int(y)}')
app.resizable(width=False, height=False)
app.configure(bg='#FFFFFF')
app.iconbitmap(sys.executable)
app.protocol("WM_DELETE_WINDOW", cerrar_ventanaPrincipal)

app.option_add("*TCombobox*Listbox*Font", font.Font(family="Helvetica", size=13))


# Frames: Contenedores dentro de la ventana que contienen los elementos
contenedor_Buscador=LabelFrame(app, text="Buscador Asunto", font=('Helvetica', 13), bg='#8297BC', fg='#FFFFFF')
contenedor_Buscador.place(x=10, y=10, width=420, height=75)
contenedor_filtros=LabelFrame(app, text="Filtros Tipo, Año y Números", font=('Helvetica', 13), bg='#C5A97B', fg='#FFFFFF')
contenedor_filtros.place(x=440, y=10, width=440, height=75)

contenedor_operaciones=LabelFrame(app, text="Operaciones", font=('Helvetica', 13), bg='#E64611', fg='#FFFFFF')
contenedor_operaciones.place(x=890, y=10, width=600, height=75)
contenedor_Tabla=Frame(app, bg='#EDB712')
contenedor_Tabla.place(x=10, y=90, width=870, height=600)
contenedor_campos=LabelFrame(app, font=('Helvetica', 13), bg='#FFFFFF', fg='#000000')
contenedor_editor=LabelFrame(app, font=('Helvetica', 13), bg='#FFFFFF', fg='#000000')
contenedor_exportar=LabelFrame(app, font=('Helvetica', 13), bg='#FFFFFF', fg='#000000')


# Tabla
global tabla
global barra1
global barra2
tabla = ttk.Treeview(contenedor_Tabla, selectmode='extended')
tabla.place(x=10, y=10, width=830, height=560)
tabla['columns']=("1", "2", "3", "4", "5", "6", "7", "8", "9")
tabla['show']='headings'
barra1=Scrollbar(contenedor_Tabla, orient=HORIZONTAL, command=tabla.xview)
barra1.place(x=10, y=575, width=830)
barra2=Scrollbar(contenedor_Tabla, orient=VERTICAL, command=tabla.yview)
barra2.place(x=845, y=10, height=560)

tabla.configure(xscrollcommand=barra1.set, yscrollcommand=barra2.set)

estilo_tabla=ttk.Style()
estilo_tabla.configure('Treeview.Heading', font=('Helvetica', 12), rowheigth=50)
estilo_tabla.configure('Treeview', font=('Helvetica', 12), rowheigth=50)
estilo_tabla.map('Treeview', background=[('selected', 'silver')])

# Encabezados de las columnas
tabla.heading('1', text="Número", anchor=W)
tabla.heading('2', text="Tipo", anchor=W)
tabla.heading('3', text="Asunto", anchor=W)
tabla.heading('4', text="Recibido de/Enviado a", anchor=W)
tabla.heading('5', text="Fecha", anchor=W)
tabla.heading('6', text="Medio", anchor=W)
tabla.heading('7', text="Número de Cheque", anchor=W)
tabla.heading('8', text="Monto", anchor=W)
tabla.heading('9', text="Por concepto de", anchor=W)

# Configuración de las columnas
tabla.column('1', stretch=NO, minwidth=80, width=80)
tabla.column('2', stretch=NO, minwidth=100, width=100)
tabla.column('3', stretch=NO, minwidth=300, width=300)
tabla.column('4', stretch=NO, minwidth=300, width=300)
tabla.column('5', stretch=NO, minwidth=100, width=100)
tabla.column('6', stretch=NO, minwidth=120, width=120)
tabla.column('7', stretch=NO, minwidth=150, width=150)
tabla.column('8', stretch=NO, minwidth=100, width=100)
tabla.column('9', stretch=NO, minwidth=800, width=800)

def deseleccionar_elemento(evento):
    tabla.selection_remove(tabla.selection())
    entradaBuscar['state']=NORMAL
    botonLimpiar['state']=NORMAL
    combo_tipo['state']='readonly'
    combo_anio['state']='readonly'
    combo_numero1['state']='readonly'
    combo_numero2['state']='readonly'
    if filtroTipo_var.get()=='Ingreso':
        botonAgregarIngreso.place(x=20, y=10)
        botonAgregarEgreso.place_forget()
    elif filtroTipo_var.get()=='Egreso':
        botonAgregarIngreso.place_forget()
        botonAgregarEgreso.place(x=180, y=10)
    else:    
        botonAgregarIngreso.place(x=20, y=10)
        botonAgregarEgreso.place(x=180, y=10)
    botonExportarExcel.place(x=340, y=10)
    botonEditar.place_forget()
    botonImprimir.place_forget()
tabla.bind("<ButtonRelease-3>", deseleccionar_elemento)

def seleccionar_elemento(evento):
    entradaBuscar['state']=DISABLED
    botonLimpiar['state']=DISABLED
    combo_tipo['state']=DISABLED
    combo_anio['state']=DISABLED
    combo_numero1['state']=DISABLED
    combo_numero2['state']=DISABLED
    botonAgregarIngreso.place_forget()
    botonAgregarEgreso.place_forget()
    botonExportarExcel.place_forget()
    botonEditar.place(x=20, y=10)
    botonImprimir.place(x=180, y=10)
tabla.bind("<ButtonRelease-1>", seleccionar_elemento)

global busqueda_var
busqueda_var=StringVar()

def habilitar_boton(evento):
    if len(entradaBuscar.get())>2:
        botonBuscar['state']=NORMAL
    else: botonBuscar['state']=DISABLED
# Entradas
global entradaBuscar
entradaBuscar=Entry(contenedor_Buscador, textvariable=busqueda_var)
entradaBuscar.place(x=20, y=10, width=205, height=32)
app.bind('<KeyRelease>', habilitar_boton)


global filtroTipo_var
global filtroA_var
global filtroNumero1_var
global filtroNumero2_var
filtroTipo_var=StringVar(value='Todos')
filtroA_var=StringVar(value='Todos')
filtroNumero1_var=StringVar(value="000")
filtroNumero2_var=StringVar(value='999')


# Combobox
combo_tipo=ttk.Combobox(contenedor_filtros, values=['Todos', 'Ingreso', 'Egreso'], textvariable=filtroTipo_var, font=("Helvetica", 12), state='readonly')
combo_tipo.place(x=10, y=10, width=100)
combo_tipo.bind('<<ComboboxSelected>>', filtrar_tabla)

lista_aFiltro=lista_a.copy()
lista_aFiltro.insert(0, 'Todos')
combo_anio=ttk.Combobox(contenedor_filtros, values=lista_aFiltro, textvariable=filtroA_var, font=("Helvetica", 12), state='readonly')
combo_anio.place(x=120, y=10, width=100)
combo_anio.bind('<<ComboboxSelected>>', filtrar_tabla)

lista_numeros1Filtro=[]
lista_numeros2Filtro=[]
for x in range(1000):
    if x>=0 and x<=9:
        lista_numeros1Filtro.append("00"+str(x))
        lista_numeros2Filtro.append("00"+str(x))
    elif x>=10 and x<=99:
        lista_numeros1Filtro.append("0"+str(x))
        lista_numeros2Filtro.append("0"+str(x))
    else:
        lista_numeros1Filtro.append(str(x))
        lista_numeros2Filtro.append(str(x))
combo_numero1=ttk.Combobox(contenedor_filtros, values=lista_numeros1Filtro, textvariable=filtroNumero1_var, font=("Helvetica", 12), state='readonly')
combo_numero1.place(x=230, y=10, width=80)
combo_numero1.bind('<<ComboboxSelected>>', filtrar_tabla)
combo_numero2=ttk.Combobox(contenedor_filtros, values=lista_numeros2Filtro, textvariable=filtroNumero2_var, font=("Helvetica", 12), state='readonly')
combo_numero2.place(x=320, y=10, width=80)
combo_numero2.bind('<<ComboboxSelected>>', filtrar_tabla)



# Botones
botonBuscar=Button(contenedor_Buscador, text="Buscar", command=buscar_asunto, font=("Helvetica", 12), bg='#FFFFFF')
botonBuscar.place(x=255, y=10)
botonBuscar['state']=DISABLED
botonLimpiar=Button(contenedor_Buscador, text="Limpiar", command=limpiar_tabla, font=("Helvetica", 12), bg='#FFFFFF')
botonLimpiar.place(x=330, y=10)
botonAgregarIngreso=Button(contenedor_operaciones, text="Agregar Ingreso", command=agregar_ingreso, font=("Helvetica", 12), bg='#FFFFFF')
botonAgregarIngreso.place(x=20, y=10)
botonAgregarEgreso=Button(contenedor_operaciones, text="Agregar Egreso", command=agregar_egreso, font=("Helvetica", 12), bg='#FFFFFF')
botonAgregarEgreso.place(x=180, y=10)
botonEditar=Button(contenedor_operaciones, text="Editar", command=editar_transaccion, font=("Helvetica", 12), bg='#FFFFFF')
botonImprimir=Button(contenedor_operaciones, text="Imprimir", command=imprimir_transaccion, font=("Helvetica", 12), bg='#FFFFFF')
botonExportarExcel=Button(contenedor_operaciones, text="Exportar a Excel", command=exportar_a_excel, font=("Helvetica", 12), bg='#FFFFFF')
botonExportarExcel.place(x=340, y=10)

# Conexión con la base de datos (importación de datos)
importar_datos_baseDeDatos()
app.mainloop()
    